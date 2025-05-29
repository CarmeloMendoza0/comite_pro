#  comite_pro/reportes/views.py
from django.shortcuts import render
from django.views import View
from transacciones.models import Transaccion, Movimiento
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.dateparse import parse_date
from catalogo_cuentas.models import Cuenta

from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class LibroDiarioView(LoginRequiredMixin, View):
    def get(self, request):
        # Filtrar transacciones por empresa y ordenar por fecha
        empresa = request.user.empresa.first()
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        #transacciones = Transaccion.objects.filter(empresa=empresa).order_by('fecha')
        transacciones = Transaccion.objects.filter(
            empresa=empresa
        ).exclude(
            comprobante__estado='Anulado'
        ).exclude(
            comprobante__activo=False  
        ).exclude(
            banco__activo=False
        ).exclude(
            activo=False
        ).order_by('fecha')

        if fecha_inicio:
            transacciones = transacciones.filter(fecha__gte=parse_date(fecha_inicio))
        if fecha_fin:
            transacciones = transacciones.filter(fecha__lte=parse_date(fecha_fin))

        # Lista para almacenar los datos del reporte
        reporte_data = []

        for transaccion in transacciones:
            movimientos = Movimiento.objects.filter(transaccion=transaccion)
            total_debe = sum(movimiento.debe or 0 for movimiento in movimientos)
            total_haber = sum(movimiento.haber or 0 for movimiento in movimientos)

            # Obtener información de tipo de transacción
            tipo_documento = ""

            # Verificar si es una transacción con documento bancario
            if transaccion.banco:
                # Obtener tipo de documento bancario
                tipo_documento = transaccion.banco.tipo_documento.nombre

            # Verificar si es una transacción con comprobante
            elif transaccion.comprobante:
                # Obtener tipo de documento comprobante
                tipo_documento = transaccion.comprobante.tipo_documento.nombre

            # Obtener la descripción de la transacción
            descripcion_transaccion = transaccion.descripcion

            # Combinar la información para mostrar en la descripción
            descripcion_completa = ""
            if tipo_documento:
                descripcion_completa += tipo_documento
            if descripcion_transaccion:
                if descripcion_completa:
                    descripcion_completa += " - "
                descripcion_completa += descripcion_transaccion
            
            # Si no se encontró ninguna información, usar la descripción original
            if not descripcion_completa:
                descripcion_completa = descripcion_transaccion

            reporte_data.append({
                'fecha': transaccion.fecha,
                'descripcion': descripcion_completa,
                'transaccion': transaccion,
                'movimientos': movimientos,
                'total_debe': total_debe,
                'total_haber': total_haber
            })

        context = {
            'reporte_data': reporte_data,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }

        return render(request, 'reportes/libro_diario.html', context)

class ExportarLibroDiarioPDFView(LoginRequiredMixin, View):
    def get(self, request):
        empresa = request.user.empresa.first()
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # Inicializar el queryset vacío
        #transacciones = Transaccion.objects.filter(empresa=empresa).order_by('fecha') if empresa else Transaccion.objects.none()
        transacciones = Transaccion.objects.filter(
            empresa=empresa
        ).exclude(
            comprobante__estado='Anulado'
        ).exclude(
            comprobante__activo=False
        ).exclude(
            banco__activo=False
        ).exclude(
            activo=False
        ).order_by('fecha') if empresa else Transaccion.objects.none()

        # Verificar si las fechas son válidas antes de filtrar
        if fecha_inicio:
            fecha_inicio = parse_date(fecha_inicio)
            if fecha_inicio:
                transacciones = transacciones.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            fecha_fin = parse_date(fecha_fin)
            if fecha_fin:
                transacciones = transacciones.filter(fecha__lte=fecha_fin)    

        # Lista para almacenar los datos del reporte
        reporte_data = []

        for transaccion in transacciones:
            movimientos = Movimiento.objects.filter(transaccion=transaccion)
            total_debe = sum(movimiento.debe or 0 for movimiento in movimientos)
            total_haber = sum(movimiento.haber or 0 for movimiento in movimientos)

            # Obtener información de tipo de transacción
            tipo_documento = ""

            # Verificar si es una transacción con documento bancario
            if transaccion.banco:
                # Obtener tipo de documento bancario
                tipo_documento = transaccion.banco.tipo_documento.nombre

            # Verificar si es una transacción con comprobante
            elif transaccion.comprobante:
                # Obtener tipo de documento comprobante
                tipo_documento = transaccion.comprobante.tipo_documento.nombre

            # Obtener la descripción de la transacción
            descripcion_transaccion = transaccion.descripcion

            # Combinar la información para mostrar en la descripción
            descripcion_completa = ""
            if tipo_documento:
                descripcion_completa += tipo_documento
            if descripcion_transaccion:
                if descripcion_completa:
                    descripcion_completa += " - "
                descripcion_completa += descripcion_transaccion
            
            # Si no se encontró ninguna información, usar la descripción original
            if not descripcion_completa:
                descripcion_completa = descripcion_transaccion

            reporte_data.append({
                'fecha': transaccion.fecha,
                'descripcion': descripcion_completa,
                'transaccion': transaccion,
                'movimientos': movimientos,
                'total_debe': total_debe,
                'total_haber': total_haber
            })

        context = {
            'reporte_data': reporte_data,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }

        # Cargar la plantilla HTML y generar el PDF
        template = get_template('reportes/libro_diario_pdf.html')
        html_content = template.render(context)
        pdf_file = HTML(string=html_content).write_pdf()

        # Responder con el archivo PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="libro_diario.pdf"'
        return response
    
class LibroMayorView(LoginRequiredMixin, View):
    def get(self, request):
        # Filtrar cuentas por empresa
        empresa = request.user.empresa.first()
        # Preservar fechas originales para el template
        fecha_inicio_str  = request.GET.get('fecha_inicio')
        fecha_fin_str  = request.GET.get('fecha_fin')
        cuenta_codigo = request.GET.get('cuenta')

        # Convertir fechas para filtros de base de datos
        fecha_inicio_parsed = None
        fecha_fin_parsed = None

        if fecha_inicio_str:
            fecha_inicio_parsed = parse_date(fecha_inicio_str)
        if fecha_fin_str:
            fecha_fin_parsed = parse_date(fecha_fin_str)

        # Filtrar cuentas de la empresa
        cuentas = Cuenta.objects.filter(catalogo__empresa=empresa)
        reporte_data = []

        # Filtrar movimientos por cuenta seleccionada o todas las cuentas
        if cuenta_codigo:
            cuentas = cuentas.filter(codigo=cuenta_codigo)

        for cuenta in cuentas:
            # Calcular el saldo acumulado hasta antes de la fecha de inicio
            saldo_anterior = cuenta.saldo
            if fecha_inicio_parsed:
                movimientos_antes = Movimiento.objects.filter(
                    cuenta=cuenta,
                    transaccion__fecha__lt=fecha_inicio_parsed
                ).exclude(
                    transaccion__comprobante__estado='Anulado'
                ).exclude(
                    transaccion__comprobante__activo=False
                ).exclude(
                    transaccion__banco__activo=False
                ).exclude(
                    transaccion__activo=False
                ).order_by('transaccion__fecha')

                # Calcular el saldo antes de la fecha de inicio
                for movimiento in movimientos_antes:
                    if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                        saldo_anterior += (movimiento.debe or 0)
                        saldo_anterior -= (movimiento.haber or 0)
                    elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                        saldo_anterior -= (movimiento.debe or 0)
                        saldo_anterior += (movimiento.haber or 0)

            # Obtener los movimientos a partir de la fecha de inicio hasta la fecha fin
            #movimientos = Movimiento.objects.filter(cuenta=cuenta).order_by('transaccion__fecha')
            movimientos = Movimiento.objects.filter(
                cuenta=cuenta
            ).exclude(
                transaccion__comprobante__estado='Anulado'
            ).exclude(
                transaccion__comprobante__activo=False
            ).exclude(
                transaccion__banco__activo=False
            ).exclude(
                transaccion__activo=False
            ).order_by('transaccion__fecha')

            if fecha_inicio_parsed:
                movimientos = movimientos.filter(transaccion__fecha__gte=fecha_inicio_parsed)
            if fecha_fin_parsed:
                movimientos = movimientos.filter(transaccion__fecha__lte=fecha_fin_parsed)

            movimientos_data = []
            total_debe = 0
            total_haber = 0
            saldo_actual = saldo_anterior

            # Calcular el saldo para cada movimiento y total de debe/haber
            for movimiento in movimientos:
                saldo_anterior_actual = saldo_actual

                if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                    saldo_actual += (movimiento.debe or 0)
                    saldo_actual -= (movimiento.haber or 0)
                elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                    saldo_actual -= (movimiento.debe or 0)
                    saldo_actual += (movimiento.haber or 0)    

                # Obtener información de tercero basado en el tipo de transacción
                nombre_tercero = ""
                tipo_documento = ""
                numero_documento = ""

                # Verificar si es una transacción con documento bancario
                if movimiento.transaccion.banco:
                    if movimiento.transaccion.banco.proveedor:
                        nombre_tercero = movimiento.transaccion.banco.proveedor.nombre
                    elif movimiento.transaccion.banco.entidad:
                        tipo_persona = movimiento.transaccion.banco.entidad.tipo
                        if tipo_persona == 'CL':
                            nombre_tercero = movimiento.transaccion.banco.entidad.nombre
                        elif tipo_persona == 'DO':
                            nombre_tercero = movimiento.transaccion.banco.entidad.nombre
                    
                    # Obtener tipo de documento bancario
                    tipo_documento = movimiento.transaccion.banco.tipo_documento.nombre
                    # Obtener número del documento bancario
                    numero_documento = movimiento.transaccion.banco.numero_documento
                
                # Verificar si es una transacción con comprobante
                elif movimiento.transaccion.comprobante:
                    if movimiento.transaccion.comprobante.proveedor:
                        nombre_tercero = movimiento.transaccion.comprobante.proveedor.nombre
                    elif movimiento.transaccion.comprobante.cliente:
                        tipo_persona = movimiento.transaccion.comprobante.cliente.tipo
                        if tipo_persona == 'CL':
                            nombre_tercero = movimiento.transaccion.comprobante.cliente.nombre
                        elif tipo_persona == 'DO':
                            nombre_tercero = movimiento.transaccion.comprobante.cliente.nombre
                    
                    # Obtener tipo de documento comprobante
                    tipo_documento = movimiento.transaccion.comprobante.tipo_documento.nombre
                    # Obtener número del documento comprobante
                    numero_documento = movimiento.transaccion.comprobante.numero_documento
                
                # Obtener la descripción de la transacción
                descripcion_transaccion = movimiento.transaccion.descripcion
                
                # Combinar la información para mostrar en la descripción
                descripcion_completa = ""
                if nombre_tercero:
                    descripcion_completa += nombre_tercero
                if tipo_documento:
                    if descripcion_completa:
                        descripcion_completa += " - "
                    descripcion_completa += tipo_documento
                if descripcion_transaccion:
                    if descripcion_completa:
                        descripcion_completa += " - "
                    descripcion_completa += descripcion_transaccion
                if numero_documento:
                    if descripcion_completa:
                        descripcion_completa += " - No. "
                    descripcion_completa += numero_documento
                
                # Si no se encontró ninguna información, usar la descripción original
                if not descripcion_completa:
                    descripcion_completa = descripcion_transaccion
                
                total_debe += (movimiento.debe or 0)
                total_haber += (movimiento.haber or 0)

                movimientos_data.append({
                    'fecha': movimiento.transaccion.fecha,
                    'descripcion': descripcion_completa,
                    'saldo_anterior': saldo_anterior_actual,
                    'debe': movimiento.debe,
                    'haber': movimiento.haber,
                    'saldo': saldo_actual
                })

            reporte_data.append({
                'codigo_cuenta': cuenta.codigo,
                'nombre_cuenta': cuenta.nombre,
                'saldo_anterior': saldo_anterior,
                'movimientos': movimientos_data,
                'total_debe': total_debe,
                'total_haber': total_haber,
                'saldo_final': saldo_actual
            })

        context = {
            'reporte_data': reporte_data,
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_fin_str,
            'cuentas': cuentas,
            'cuenta_seleccionada': cuenta_codigo,
        }

        return render(request, 'reportes/libro_mayor.html', context)
    
class ExportarLibroMayorPDFView(LoginRequiredMixin, View):
    def get(self, request):
        # Filtrar cuentas por empresa
        empresa = request.user.empresa.first()
        # Preservar fechas originales para el template
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        cuenta_codigo = request.GET.get('cuenta')

        # Convertir fechas para filtros de base de datos
        fecha_inicio_parsed = None
        fecha_fin_parsed = None
        
        if fecha_inicio_str:
            fecha_inicio_parsed = parse_date(fecha_inicio_str)
        if fecha_fin_str:
            fecha_fin_parsed = parse_date(fecha_fin_str)

        # Filtrar cuentas de la empresa
        cuentas = Cuenta.objects.filter(catalogo__empresa=empresa)
        # Lista para almacenar los datos del reporte
        reporte_data = []

        # Filtrar movimientos por cuenta seleccionada o todas las cuentas
        if cuenta_codigo:
            cuentas = cuentas.filter(codigo=cuenta_codigo)

        for cuenta in cuentas:
            # Calcular el saldo acumulado hasta antes de la fecha de inicio
            saldo_anterior = cuenta.saldo
            if fecha_inicio_parsed:
                movimientos_antes = Movimiento.objects.filter(
                    cuenta=cuenta,
                    transaccion__fecha__lt=fecha_inicio_parsed
                ).exclude(
                    transaccion__comprobante__estado='Anulado'
                ).exclude(
                    transaccion__comprobante__activo=False
                ).exclude(
                    transaccion__banco__activo=False
                ).exclude(
                    transaccion__activo=False
                ).order_by('transaccion__fecha')

                # Calcular el saldo antes de la fecha de inicio
                for movimiento in movimientos_antes:
                    if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                        saldo_anterior += (movimiento.debe or 0)
                        saldo_anterior -= (movimiento.haber or 0)
                    elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                        saldo_anterior -= (movimiento.debe or 0)
                        saldo_anterior += (movimiento.haber or 0)

            # Obtener los movimientos a partir de la fecha de inicio hasta la fecha fin
            #movimientos = Movimiento.objects.filter(cuenta=cuenta).order_by('transaccion__fecha')
            movimientos = Movimiento.objects.filter(
                cuenta=cuenta
            ).exclude(
                transaccion__comprobante__estado='Anulado'
            ).exclude(
                transaccion__comprobante__activo=False
            ).exclude(
                transaccion__banco__activo=False
            ).exclude(
                transaccion__activo=False
            ).order_by('transaccion__fecha')

            if fecha_inicio_parsed:
                movimientos = movimientos.filter(transaccion__fecha__gte=fecha_inicio_parsed)
            if fecha_fin_parsed:
                movimientos = movimientos.filter(transaccion__fecha__lte=fecha_fin_parsed)

            movimientos_data = []
            total_debe = 0
            total_haber = 0
            saldo_actual = saldo_anterior

            # Calcular el saldo para cada movimiento y total de debe/haber
            for movimiento in movimientos:
                saldo_anterior_actual = saldo_actual
                if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                    saldo_actual += (movimiento.debe or 0)
                    saldo_actual -= (movimiento.haber or 0)
                elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                    saldo_actual -= (movimiento.debe or 0)
                    saldo_actual += (movimiento.haber or 0)

                # Obtener información de tercero basado en el tipo de transacción
                nombre_tercero = ""
                tipo_documento = ""
                numero_documento = ""

                # Verificar si es una transacción con documento bancario
                if movimiento.transaccion.banco:
                    if movimiento.transaccion.banco.proveedor:
                        nombre_tercero = movimiento.transaccion.banco.proveedor.nombre
                    elif movimiento.transaccion.banco.entidad:
                        tipo_persona = movimiento.transaccion.banco.entidad.tipo
                        if tipo_persona == 'CL':
                            nombre_tercero = movimiento.transaccion.banco.entidad.nombre
                        elif tipo_persona == 'DO':
                            nombre_tercero = movimiento.transaccion.banco.entidad.nombre
                    
                    # Obtener tipo de documento bancario
                    tipo_documento = movimiento.transaccion.banco.tipo_documento.nombre
                    # Obtener número del documento bancario
                    numero_documento = movimiento.transaccion.banco.numero_documento
                
                # Verificar si es una transacción con comprobante
                elif movimiento.transaccion.comprobante:
                    if movimiento.transaccion.comprobante.proveedor:
                        nombre_tercero = movimiento.transaccion.comprobante.proveedor.nombre
                    elif movimiento.transaccion.comprobante.cliente:
                        tipo_persona = movimiento.transaccion.comprobante.cliente.tipo
                        if tipo_persona == 'CL':
                            nombre_tercero = movimiento.transaccion.comprobante.cliente.nombre
                        elif tipo_persona == 'DO':
                            nombre_tercero = movimiento.transaccion.comprobante.cliente.nombre
                    
                    # Obtener tipo de documento comprobante
                    tipo_documento = movimiento.transaccion.comprobante.tipo_documento.nombre
                    # Obtener número del documento comprobante
                    numero_documento = movimiento.transaccion.comprobante.numero_documento
                
                # Obtener la descripción de la transacción
                descripcion_transaccion = movimiento.transaccion.descripcion
                
                # Combinar la información para mostrar en la descripción
                descripcion_completa = ""
                if nombre_tercero:
                    descripcion_completa += nombre_tercero
                if tipo_documento:
                    if descripcion_completa:
                        descripcion_completa += " - "
                    descripcion_completa += tipo_documento
                if descripcion_transaccion:
                    if descripcion_completa:
                        descripcion_completa += " - "
                    descripcion_completa += descripcion_transaccion
                if numero_documento:
                    if descripcion_completa:
                        descripcion_completa += " - No. "
                    descripcion_completa += numero_documento
                
                # Si no se encontró ninguna información, usar la descripción original
                if not descripcion_completa:
                    descripcion_completa = descripcion_transaccion

                total_debe += (movimiento.debe or 0)
                total_haber += (movimiento.haber or 0)

                movimientos_data.append({
                    'fecha': movimiento.transaccion.fecha,
                    'descripcion': descripcion_completa,
                    'saldo_anterior': saldo_anterior_actual,
                    'debe': movimiento.debe,
                    'haber': movimiento.haber,
                    'saldo': saldo_actual
                })

            reporte_data.append({
                'codigo_cuenta': cuenta.codigo,
                'nombre_cuenta': cuenta.nombre,
                'saldo_anterior': saldo_anterior,
                'movimientos': movimientos_data,
                'total_debe': total_debe,
                'total_haber': total_haber,
                'saldo_final': saldo_actual
            })

        context = {
            'reporte_data': reporte_data,
            'fecha_inicio': fecha_inicio_parsed,
            'fecha_fin': fecha_fin_parsed,
            'cuentas': cuentas,
            'cuenta_seleccionada': cuenta_codigo,
        }

        # Cargar la plantilla HTML y generar el PDF
        template = get_template('reportes/libro_mayor_pdf.html')
        html_content = template.render(context)
        pdf_file = HTML(string=html_content).write_pdf()

        # Responder con el archivo PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="libro_mayor.pdf"'
        return response
    
class ExportarLibroDiarioExcelView(LoginRequiredMixin, View):
    def get(self, request):
        # Utilizar la misma lógica de filtro que la vista LibroDiarioView
        empresa = request.user.empresa.first()
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # Inicializar el queryset
        #transacciones = Transaccion.objects.filter(empresa=empresa).order_by('fecha') if empresa else Transaccion.objects.none()
        transacciones = Transaccion.objects.filter(
            empresa=empresa
        ).exclude(
            comprobante__estado='Anulado'
        ).exclude(
            comprobante__activo=False  
        ).exclude(
            banco__activo=False
        ).exclude(
            activo=False
        ).order_by('fecha') if empresa else Transaccion.objects.none()

        # Aplicar filtros de fecha
        fecha_inicio_parsed = None
        fecha_fin_parsed = None
        
        if fecha_inicio and fecha_inicio != 'None':
            fecha_inicio_parsed = parse_date(fecha_inicio)
            if fecha_inicio_parsed:
                transacciones = transacciones.filter(fecha__gte=fecha_inicio_parsed)
        if fecha_fin and fecha_fin != 'None':
            fecha_fin_parsed = parse_date(fecha_fin)
            if fecha_fin_parsed:
                transacciones = transacciones.filter(fecha__lte=fecha_fin_parsed)

        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"
        
        # Configurar encabezados del reporte
        ws['A1'] = "LIBRO DIARIO - COMITÉ PRO-OPERACIÓN"
        
        # Fechas del período
        if fecha_inicio_parsed and fecha_fin_parsed:
            ws['A2'] = f"Del {fecha_inicio_parsed.strftime('%Y-%m-%d')} al {fecha_fin_parsed.strftime('%Y-%m-%d')}"
        else:
            ws['A2'] = "Todos los registros"
        
        # Combinar celdas para el título
        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')
        
        # Aplicar formato al título
        title_font = Font(size=14, bold=True)
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Agregar encabezados de columnas
        headers = ['Fecha/Código', 'Cuenta/Partida-Descripción', 'Debe', 'Haber']
        header_row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Inicializar fila actual
        current_row = header_row + 1
        partida_num = 1
        
        # Variables para totales generales
        total_debe_general = 0
        total_haber_general = 0
        
        # Procesar transacciones
        for transaccion in transacciones:
            movimientos = Movimiento.objects.filter(transaccion=transaccion).select_related('cuenta')
            
            if not movimientos:
                continue
                
            total_debe = sum(float(movimiento.debe or 0) for movimiento in movimientos)
            total_haber = sum(float(movimiento.haber or 0) for movimiento in movimientos)

            # Obtener información de tipo de transacción
            tipo_documento = ""
            if hasattr(transaccion, 'banco') and transaccion.banco:
                tipo_documento = transaccion.banco.tipo_documento.nombre
            elif hasattr(transaccion, 'comprobante') and transaccion.comprobante:
                tipo_documento = transaccion.comprobante.tipo_documento.nombre

            # Obtener la descripción completa
            descripcion_transaccion = transaccion.descripcion or ""
            descripcion_completa = ""
            
            if tipo_documento:
                descripcion_completa += tipo_documento
            if descripcion_transaccion:
                if descripcion_completa:
                    descripcion_completa += " - "
                descripcion_completa += descripcion_transaccion
            
            if not descripcion_completa:
                descripcion_completa = "Sin descripción"

            # Escribir fecha
            ws.cell(row=current_row, column=1, value=transaccion.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            
            # Escribir encabezado de partida
            ws.cell(row=current_row, column=2, value=f'Partida #{partida_num} {descripcion_completa}')
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
            # Aplicar bordes a la fila de partida
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            current_row += 1
            
            # Escribir movimientos
            for movimiento in movimientos:
                # Código de cuenta
                ws.cell(row=current_row, column=1, value=movimiento.cuenta.codigo)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
                
                # Nombre de cuenta
                ws.cell(row=current_row, column=2, value=movimiento.cuenta.nombre)
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')
                
                # Debe
                debe_value = float(movimiento.debe) if movimiento.debe else 0
                if debe_value > 0:
                    ws.cell(row=current_row, column=3, value=debe_value)
                    ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
                    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right')
                
                # Haber
                haber_value = float(movimiento.haber) if movimiento.haber else 0
                if haber_value > 0:
                    ws.cell(row=current_row, column=4, value=haber_value)
                    ws.cell(row=current_row, column=4).number_format = '"Q"#,##0.00'
                    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right')
                
                # Aplicar bordes
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                current_row += 1
            
            ws.cell(row=current_row, column=2, value='TOTAL')
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
            
            ws.cell(row=current_row, column=3, value=total_debe)
            ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=3).font = Font(bold=True)
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right')
            
            ws.cell(row=current_row, column=4, value=total_haber)
            ws.cell(row=current_row, column=4).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=4).font = Font(bold=True)
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right')
            
            # Formato para la fila de totales
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='double'),
                    bottom=Side(style='double')
                )
            
            # Actualizar totales generales
            total_debe_general += total_debe
            total_haber_general += total_haber
            
            current_row += 2  # Dejar una fila vacía
            partida_num += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15  # Fecha/Código
        ws.column_dimensions['B'].width = 50  # Cuenta/Descripción
        ws.column_dimensions['C'].width = 15  # Debe
        ws.column_dimensions['D'].width = 15  # Haber
        
        # Agregar resumen al final
        current_row += 2
        ws.cell(row=current_row, column=1, value="RESUMEN DEL PERÍODO")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 1
        ws.cell(row=current_row, column=2, value="Total Debe:")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=3, value=total_debe_general)
        ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        
        current_row += 1
        ws.cell(row=current_row, column=2, value="Total Haber:")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=3, value=total_haber_general)
        ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        
        current_row += 1
        ws.cell(row=current_row, column=2, value="Diferencia:")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=3, value=abs(total_debe_general - total_haber_general))
        ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        
        # Aplicar bordes al resumen
        for row in range(current_row - 3, current_row + 1):
            for col in range(2, 4):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
        
        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo con fechas si están disponibles
        if fecha_inicio_parsed and fecha_fin_parsed:
            filename = f"libro_diario_{fecha_inicio_parsed.strftime('%Y%m%d')}_{fecha_fin_parsed.strftime('%Y%m%d')}.xlsx"
        else:
            filename = "libro_diario.xlsx"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el archivo
        wb.save(response)
        return response
    
class ExportarLibroMayorExcelView(LoginRequiredMixin, View):
    def get(self, request):
        # Filtrar cuentas por empresa
        empresa = request.user.empresa.first()
        # Preservar fechas originales
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        cuenta_codigo = request.GET.get('cuenta')

        # Convertir fechas para filtros de base de datos
        fecha_inicio_parsed = None
        fecha_fin_parsed = None
        
        if fecha_inicio_str and fecha_inicio_str != 'None':
            fecha_inicio_parsed = parse_date(fecha_inicio_str)
        if fecha_fin_str and fecha_fin_str != 'None':
            fecha_fin_parsed = parse_date(fecha_fin_str)

        # Filtrar cuentas de la empresa
        cuentas = Cuenta.objects.filter(catalogo__empresa=empresa) if empresa else Cuenta.objects.none()
        
        # Filtrar por cuenta específica si se seleccionó
        if cuenta_codigo:
            cuentas = cuentas.filter(codigo=cuenta_codigo)

        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Mayor"
        
        # Configurar encabezados del reporte
        ws['A1'] = "COMITÉ PRO-OPERACIÓN"
        ws['A2'] = "LIBRO MAYOR"
        
        # Período
        periodo_texto = "Período: "
        if fecha_inicio_parsed and fecha_fin_parsed:
            periodo_texto += f"{fecha_inicio_parsed.strftime('%d/%m/%Y')} al {fecha_fin_parsed.strftime('%d/%m/%Y')}"
        elif fecha_inicio_parsed:
            periodo_texto += f"{fecha_inicio_parsed.strftime('%d/%m/%Y')} al Actual"
        elif fecha_fin_parsed:
            periodo_texto += f"Inicio al {fecha_fin_parsed.strftime('%d/%m/%Y')}"
        else:
            periodo_texto += "Todos los registros"
        ws['A3'] = periodo_texto
        
        # Información de cuenta seleccionada
        if cuenta_codigo:
            cuenta_info = cuentas.first()
            if cuenta_info:
                ws['A4'] = f"Cuenta: {cuenta_info.codigo} - {cuenta_info.nombre}"
        else:
            ws['A4'] = "Todas las Cuentas"
        
        # Combinar celdas para el título
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        
        # Aplicar formato al título
        for row in range(1, 5):
            cell = ws[f'A{row}']
            cell.font = Font(size=14 if row == 1 else 12, bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Agregar encabezados de columnas
        headers = ['Cuenta / Fecha', 'Descripción', 'Saldo Anterior', 'Debe', 'Haber', 'Saldo']
        header_row = 6
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Inicializar fila actual
        current_row = header_row + 1
        
        # Procesar cada cuenta
        for cuenta in cuentas:
            # Calcular el saldo acumulado hasta antes de la fecha de inicio
            saldo_anterior = float(cuenta.saldo or 0)
            
            if fecha_inicio_parsed:
                movimientos_antes = Movimiento.objects.filter(
                    cuenta=cuenta,
                    transaccion__fecha__lt=fecha_inicio_parsed
                ).exclude(
                    transaccion__comprobante__estado='Anulado'
                ).exclude(
                    transaccion__comprobante__activo=False
                ).exclude(
                    transaccion__banco__activo=False
                ).exclude(
                    transaccion__activo=False
                ).order_by('transaccion__fecha')

                # Calcular el saldo antes de la fecha de inicio
                for movimiento in movimientos_antes:
                    if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                        saldo_anterior += float(movimiento.debe or 0)
                        saldo_anterior -= float(movimiento.haber or 0)
                    elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                        saldo_anterior -= float(movimiento.debe or 0)
                        saldo_anterior += float(movimiento.haber or 0)

            # Obtener los movimientos del período
            #movimientos = Movimiento.objects.filter(cuenta=cuenta).order_by('transaccion__fecha')
            movimientos = Movimiento.objects.filter(
                cuenta=cuenta
            ).exclude(
                transaccion__comprobante__estado='Anulado'
            ).exclude(
                transaccion__comprobante__activo=False
            ).exclude(
                transaccion__banco__activo=False
            ).exclude(
                transaccion__activo=False
            ).order_by('transaccion__fecha')

            if fecha_inicio_parsed:
                movimientos = movimientos.filter(transaccion__fecha__gte=fecha_inicio_parsed)
            if fecha_fin_parsed:
                movimientos = movimientos.filter(transaccion__fecha__lte=fecha_fin_parsed)

            # Si no hay movimientos y no se especificó una cuenta, continuar
            if not movimientos and not cuenta_codigo:
                continue

            # Escribir encabezado de cuenta
            ws.cell(row=current_row, column=1, value=cuenta.codigo)
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="2c3e50")
            
            ws.cell(row=current_row, column=2, value=cuenta.nombre)
            ws.cell(row=current_row, column=2).font = Font(bold=True, color="34495e")
            
            ws.cell(row=current_row, column=3, value=saldo_anterior)
            ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=3).font = Font(bold=True)
            
            # Aplicar formato a la fila de encabezado de cuenta
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')
                )
            
            current_row += 1
            
            total_debe = 0
            total_haber = 0
            saldo_actual = saldo_anterior
            
            # Escribir movimientos
            for idx, movimiento in enumerate(movimientos):
                # Fecha
                ws.cell(row=current_row, column=1, value=movimiento.transaccion.fecha.strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
                
                # Descripción
                descripcion = self._obtener_descripcion_completa(movimiento)
                ws.cell(row=current_row, column=2, value=descripcion[:100])  # Limitar a 100 caracteres
                
                # Saldo anterior
                ws.cell(row=current_row, column=3, value=saldo_actual)
                ws.cell(row=current_row, column=3).number_format = '"Q"#,##0.00'
                
                # Debe
                debe_value = float(movimiento.debe or 0)
                if debe_value > 0:
                    ws.cell(row=current_row, column=4, value=debe_value)
                    ws.cell(row=current_row, column=4).number_format = '"Q"#,##0.00'
                total_debe += debe_value
                
                # Haber
                haber_value = float(movimiento.haber or 0)
                if haber_value > 0:
                    ws.cell(row=current_row, column=5, value=haber_value)
                    ws.cell(row=current_row, column=5).number_format = '"Q"#,##0.00'
                total_haber += haber_value
                
                # Calcular nuevo saldo
                if cuenta.catalogo.tipo in ['Activo', 'Gasto']:  # Naturaleza deudora
                    saldo_actual += debe_value - haber_value
                elif cuenta.catalogo.tipo in ['Pasivo', 'Capital', 'Ingreso']:  # Naturaleza acreedora
                    saldo_actual += haber_value - debe_value
                
                # Saldo
                ws.cell(row=current_row, column=6, value=saldo_actual)
                ws.cell(row=current_row, column=6).number_format = '"Q"#,##0.00'
                ws.cell(row=current_row, column=6).font = Font(bold=True)
                
                # Color para saldos negativos
                if saldo_actual < 0:
                    ws.cell(row=current_row, column=6).font = Font(bold=True, color="e74c3c")
                else:
                    ws.cell(row=current_row, column=6).font = Font(bold=True, color="27ae60")
                
                # Aplicar bordes y fondo alternado
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    if idx % 2 == 1:
                        cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
                
                current_row += 1
            
            # Si no hay movimientos, mostrar mensaje
            if not movimientos:
                ws.cell(row=current_row, column=1, value="Sin movimientos en el período seleccionado")
                ws.merge_cells(f'A{current_row}:F{current_row}')
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row, column=1).font = Font(italic=True)
                current_row += 1
            
            # Escribir totales
            ws.cell(row=current_row, column=1, value='TOTALES')
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            ws.cell(row=current_row, column=2, value=cuenta.nombre)
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
            ws.cell(row=current_row, column=4, value=total_debe)
            ws.cell(row=current_row, column=4).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=4).font = Font(bold=True)
            
            ws.cell(row=current_row, column=5, value=total_haber)
            ws.cell(row=current_row, column=5).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=5).font = Font(bold=True)
            
            ws.cell(row=current_row, column=6, value=saldo_actual)
            ws.cell(row=current_row, column=6).number_format = '"Q"#,##0.00'
            ws.cell(row=current_row, column=6).font = Font(bold=True)
            
            # Color para saldo final
            if saldo_actual < 0:
                ws.cell(row=current_row, column=6).font = Font(bold=True, color="e74c3c")
            else:
                ws.cell(row=current_row, column=6).font = Font(bold=True, color="27ae60")
            
            # Formato para la fila de totales
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')
                )
            
            current_row += 2  # Espacio entre cuentas
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 18  # Cuenta/Fecha
        ws.column_dimensions['B'].width = 45  # Descripción
        ws.column_dimensions['C'].width = 18  # Saldo Anterior
        ws.column_dimensions['D'].width = 15  # Debe
        ws.column_dimensions['E'].width = 15  # Haber
        ws.column_dimensions['F'].width = 18  # Saldo
        
        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo
        if fecha_inicio_parsed and fecha_fin_parsed:
            filename = f"libro_mayor_{fecha_inicio_parsed.strftime('%Y%m%d')}_{fecha_fin_parsed.strftime('%Y%m%d')}"
        else:
            filename = "libro_mayor"
            
        if cuenta_codigo:
            filename += f"_{cuenta_codigo}"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        # Guardar el archivo
        wb.save(response)
        return response
    
    def _obtener_descripcion_completa(self, movimiento):
        """Método auxiliar para obtener la descripción completa del movimiento"""
        nombre_tercero = ""
        tipo_documento = ""
        numero_documento = ""

        # Verificar si es una transacción con documento bancario
        if hasattr(movimiento.transaccion, 'banco') and movimiento.transaccion.banco:
            if movimiento.transaccion.banco.proveedor:
                nombre_tercero = movimiento.transaccion.banco.proveedor.nombre
            elif movimiento.transaccion.banco.entidad:
                nombre_tercero = movimiento.transaccion.banco.entidad.nombre
            
            tipo_documento = movimiento.transaccion.banco.tipo_documento.nombre
            numero_documento = movimiento.transaccion.banco.numero_documento
        
        # Verificar si es una transacción con comprobante
        elif hasattr(movimiento.transaccion, 'comprobante') and movimiento.transaccion.comprobante:
            if movimiento.transaccion.comprobante.proveedor:
                nombre_tercero = movimiento.transaccion.comprobante.proveedor.nombre
            elif movimiento.transaccion.comprobante.cliente:
                nombre_tercero = movimiento.transaccion.comprobante.cliente.nombre
            
            tipo_documento = movimiento.transaccion.comprobante.tipo_documento.nombre
            numero_documento = movimiento.transaccion.comprobante.numero_documento
        
        # Obtener la descripción de la transacción
        descripcion_transaccion = movimiento.transaccion.descripcion or ""
        
        # Combinar la información
        descripcion_completa = ""
        if nombre_tercero:
            descripcion_completa += nombre_tercero
        if tipo_documento:
            if descripcion_completa:
                descripcion_completa += " - "
            descripcion_completa += tipo_documento
        if descripcion_transaccion:
            if descripcion_completa:
                descripcion_completa += " - "
            descripcion_completa += descripcion_transaccion
        if numero_documento:
            if descripcion_completa:
                descripcion_completa += " - No. "
            descripcion_completa += numero_documento
        
        return descripcion_completa or "Sin descripción"


